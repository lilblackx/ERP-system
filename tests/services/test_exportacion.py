from openpyxl import load_workbook

from app.services.exportacion import exportar_excel


def test_exportar_excel_escribe_encabezados_y_filas(tmp_path):
    ruta = tmp_path / "reporte.xlsx"

    exportar_excel(ruta, ["Cliente", "Saldo"], [["Cliente A", 100], ["Cliente B", 50]])

    libro = load_workbook(ruta)
    hoja = libro.active
    filas = [[celda.value for celda in fila] for fila in hoja.iter_rows()]
    assert filas == [["Cliente", "Saldo"], ["Cliente A", 100], ["Cliente B", 50]]


def test_exportar_excel_sin_filas_escribe_solo_encabezados(tmp_path):
    ruta = tmp_path / "vacio.xlsx"

    exportar_excel(ruta, ["Cliente", "Saldo"], [])

    libro = load_workbook(ruta)
    hoja = libro.active
    filas = [[celda.value for celda in fila] for fila in hoja.iter_rows()]
    assert filas == [["Cliente", "Saldo"]]


def test_exportar_excel_acepta_generador_de_filas(tmp_path):
    ruta = tmp_path / "generador.xlsx"

    def filas():
        yield ["A", 1]
        yield ["B", 2]

    exportar_excel(ruta, ["Letra", "Numero"], filas())

    libro = load_workbook(ruta)
    hoja = libro.active
    valores = [[celda.value for celda in fila] for fila in hoja.iter_rows()]
    assert valores == [["Letra", "Numero"], ["A", 1], ["B", 2]]
