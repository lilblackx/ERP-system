"""R-02 (docs/CHECKLIST_PRODUCCION.md): exportacion de reportes/listados a Excel.

Escritor generico (encabezados + filas) reutilizado por el motor de reportes
(app/services/reportes.py) y por listados de catalogo con boton "Exportar" (ej.
app/ui/clientes_panel.py, R-10).

R-09: la ruta de destino la elige el caller (tipicamente via
QFileDialog.getSaveFileName() en la UI) ANTES de llamar a esta funcion -- se escribe
directo ahi, nunca a un archivo temporal, asi que no hace falta purgar nada al cerrar
la app.

No usa `write_only=True` (streaming): eso es para datasets grandes que no caben comodos
en memoria (kardex completo, facturas del año -- ver R-04, pendiente hasta que exista un
reporte de ese tamaño). Los listados actuales (aging de CxC, catalogo de clientes) son
acotados.
"""

import logging
from collections.abc import Iterable, Sequence
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.utils import ImageReader
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.db.models import ConfiguracionEmpresa

logger = logging.getLogger(__name__)

_COLOR_PRIMARIO = colors.HexColor("#0D47A1")
_COLOR_MUTED = colors.HexColor("#64748B")


def exportar_excel(
    ruta: str | Path,
    encabezados: Sequence[str],
    filas: Iterable[Sequence[Any]],
    titulo: str | None = None,
    config_empresa: ConfiguracionEmpresa | None = None,
) -> None:
    """Escribe una unica hoja (encabezados + filas) a un archivo .xlsx en `ruta`.

    `titulo`/`config_empresa` son opcionales para no romper a los callers existentes
    (catalogos de clientes/bancos/inventario/etc., que solo pasan encabezados+filas): sin
    ellos el archivo queda igual que antes (solo encabezados+filas desde la fila 1). Con
    `config_empresa` se antepone la razon social/RIF/direccion/telefono de la empresa
    (sin logo -- Excel no lo necesita) y, si hay `titulo`, el nombre del reporte centrado
    una fila en blanco despues de esos datos -- pedido del usuario 2026-09-01 para que
    todo reporte exportado (Excel o PDF) traiga la info de la empresa."""
    libro = Workbook()
    hoja = libro.active
    ancho_tabla = max(len(encabezados), 1)
    fila_actual = 1

    if config_empresa is not None:
        lineas_empresa = (
            (config_empresa.razon_social_empresa, Font(bold=True, size=13, color="0D47A1")),
            (config_empresa.rif_empresa, Font(size=10, color="64748B")),
            (config_empresa.direccion_empresa, Font(size=10, color="64748B")),
            (config_empresa.telefono_empresa, Font(size=10, color="64748B")),
        )
        escribio_alguna = False
        for indice, (valor, fuente) in enumerate(lineas_empresa):
            if not valor:
                continue
            celda = hoja.cell(row=fila_actual, column=1, value=valor)
            celda.font = fuente
            if indice == 0:
                # mas alto que el resto -- separa visualmente la razon social del RIF que
                # sigue justo debajo, quedaban muy pegados (hallazgo del usuario).
                hoja.row_dimensions[fila_actual].height = 22
            fila_actual += 1
            escribio_alguna = True
        if escribio_alguna:
            fila_actual += 1  # fila en blanco antes del titulo

    if titulo:
        celda_titulo = hoja.cell(row=fila_actual, column=1, value=titulo)
        celda_titulo.font = Font(bold=True, size=14)
        celda_titulo.alignment = Alignment(horizontal="center")
        if ancho_tabla > 1:
            hoja.merge_cells(start_row=fila_actual, start_column=1, end_row=fila_actual, end_column=ancho_tabla)
        fila_actual += 2  # fila en blanco despues del titulo

    for col_idx, encabezado in enumerate(encabezados, start=1):
        hoja.cell(row=fila_actual, column=col_idx, value=encabezado).font = Font(bold=True)
    fila_actual += 1

    cantidad_filas = 0
    for fila in filas:
        for col_idx, valor in enumerate(fila, start=1):
            hoja.cell(row=fila_actual, column=col_idx, value=valor)
        fila_actual += 1
        cantidad_filas += 1

    libro.save(ruta)
    logger.info("Excel exportado: %s (%d filas)", ruta, cantidad_filas)


def _imagen_logo_pdf(logo_bytes: bytes, dimension_max: float = 55) -> Image | None:
    """Escala el logo (bytes PNG/JPG guardados en ConfiguracionEmpresa.logotipo_empresa,
    subidos sin control de tamaño en config_empresa_panel.py) a un cuadro de a lo sumo
    `dimension_max` puntos, preservando proporcion -- sin esto reportlab dibujaria la
    imagen a su tamaño nativo en pixeles interpretado como puntos, facilmente mas grande
    que la pagina."""
    try:
        ancho_px, alto_px = ImageReader(BytesIO(logo_bytes)).getSize()
    except (OSError, ValueError):
        return None
    if not ancho_px or not alto_px:
        return None
    escala = dimension_max / max(ancho_px, alto_px)
    return Image(BytesIO(logo_bytes), width=ancho_px * escala, height=alto_px * escala)


def _bloque_empresa_pdf(config_empresa: ConfiguracionEmpresa, styles) -> Table | Paragraph | None:
    estilo_nombre = ParagraphStyle(
        "EmpresaNombre",
        parent=styles["Normal"],
        fontSize=13,
        fontName="Helvetica-Bold",
        textColor=_COLOR_PRIMARIO,
        spaceAfter=5,
    )
    estilo_dato = ParagraphStyle("EmpresaDato", parent=styles["Normal"], fontSize=9, textColor=_COLOR_MUTED)

    lineas: list[Paragraph] = []
    if config_empresa.razon_social_empresa:
        lineas.append(Paragraph(config_empresa.razon_social_empresa, estilo_nombre))
    for valor in (config_empresa.rif_empresa, config_empresa.direccion_empresa, config_empresa.telefono_empresa):
        if valor:
            lineas.append(Paragraph(valor, estilo_dato))
    if not lineas:
        return None

    logo = _imagen_logo_pdf(config_empresa.logotipo_empresa) if config_empresa.logotipo_empresa else None
    if logo is None:
        return (
            lineas[0]
            if len(lineas) == 1
            else Table([[linea] for linea in lineas], style=[("LEFTPADDING", (0, 0), (-1, -1), 0)])
        )

    fila = Table([[logo, lineas]], colWidths=[65, None])
    fila.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (0, 0), 0),
                ("LEFTPADDING", (1, 0), (1, 0), 10),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )
    return fila


def exportar_pdf(
    ruta: str | Path,
    titulo: str,
    encabezados: Sequence[str],
    filas: Iterable[Sequence[Any]],
    cliente_nombre: str | None = None,
    filtros: dict[str, Any] | None = None,
    col_widths: Sequence[float] | None = None,
    config_empresa: ConfiguracionEmpresa | None = None,
) -> None:
    """Exporta datos a un archivo PDF con formato de tabla.

    Args:
        ruta: Ruta del archivo PDF a generar.
        titulo: Título principal del reporte.
        encabezados: Lista de nombres de columnas.
        filas: Filas de datos a exportar.
        cliente_nombre: Nombre del cliente (opcional, para reportes específicos).
        filtros: Diccionario con los filtros aplicados (opcional).
        col_widths: Anchos específicos para cada columna (opcional).
        config_empresa: datos de la empresa (razón social, RIF, dirección, teléfono,
            logo) a imprimir arriba del título -- opcional para no romper a los callers
            existentes que no lo pasan (catálogos de clientes/bancos/inventario/etc.).
    """
    margen_horizontal = 30
    doc = SimpleDocTemplate(
        str(ruta),
        pagesize=letter,
        rightMargin=margen_horizontal,
        leftMargin=margen_horizontal,
        topMargin=30,
        bottomMargin=18,
    )

    elements = []
    styles = getSampleStyleSheet()

    # Datos de la empresa (con logo) arriba de todo -- pedido del usuario 2026-09-01.
    if config_empresa is not None:
        bloque_empresa = _bloque_empresa_pdf(config_empresa, styles)
        if bloque_empresa is not None:
            elements.append(bloque_empresa)
            elements.append(Spacer(1, 14))  # 1-2 lineas de separacion antes del titulo

    # Título del reporte con fecha y hora, centrado
    fecha_hora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    titulo_completo = f"{titulo} - {fecha_hora}"
    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Heading1"],
        fontSize=16,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#0D47A1"),
        spaceAfter=12,
    )
    elements.append(Paragraph(titulo_completo, title_style))

    # Información de filtros aplicados
    if filtros:
        filtro_parts = []
        for key, value in filtros.items():
            if value:
                filtro_parts.append(f"{key}: {value}")
        if filtro_parts:
            filtro_texto = " | ".join(filtro_parts)
            filtro_style = ParagraphStyle(
                "Filtros",
                parent=styles["Normal"],
                fontSize=10,
                textColor=colors.HexColor("#64748B"),
                spaceAfter=12,
            )
            elements.append(Paragraph(f"Filtros: {filtro_texto}", filtro_style))

    # Nombre del cliente si se proporciona
    if cliente_nombre:
        cliente_style = ParagraphStyle(
            "Cliente",
            parent=styles["Normal"],
            fontSize=12,
            textColor=colors.HexColor("#64748B"),
            spaceAfter=12,
        )
        elements.append(Paragraph(f"Cliente: {cliente_nombre}", cliente_style))

    # Datos de la tabla
    data = [list(encabezados)]
    for fila in filas:
        data.append(list(fila))

    # Estilo de la tabla
    table_style = TableStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0D47A1")),  # Header azul
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ("TOPPADDING", (0, 0), (-1, 0), 8),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("WORDWRAP", (0, 0), (-1, -1), "CJK"),  # Permite mejor ajuste de texto largo
        ]
    )

    # Alternar colores de filas
    for i in range(1, len(data)):
        if i % 2 == 0:
            table_style.add("BACKGROUND", (0, i), (-1, i), colors.HexColor("#F8FAFC"))
        else:
            table_style.add("BACKGROUND", (0, i), (-1, i), colors.white)

    # Ancho de columnas: usar anchos específicos si se proporcionan,
    # de lo contrario repartir equitativamente
    ancho_disponible = letter[0] - 2 * margen_horizontal
    if col_widths:
        # Normalizar anchos para que sumen el ancho disponible
        total_ancho = sum(col_widths)
        col_widths_normalizados = [w * ancho_disponible / total_ancho for w in col_widths]
        table = Table(data, colWidths=col_widths_normalizados)
    else:
        ancho_columna = ancho_disponible / len(encabezados)
        table = Table(data, colWidths=[ancho_columna] * len(encabezados))
    table.setStyle(table_style)
    elements.append(table)

    doc.build(elements)
    logger.info("PDF exportado: %s", ruta)
